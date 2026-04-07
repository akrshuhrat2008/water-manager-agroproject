"""API для импорта и экспорта данных."""
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO, StringIO
import json
import csv
from typing import List, Dict, Any

router = APIRouter()


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """Импортировать данные из Excel."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть Excel (.xlsx или .xls)")
    
    try:
        from openpyxl import load_workbook
        
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Читаем данные из Excel
        data = []
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(cell) if cell else f"Column{idx+1}" for idx, cell in enumerate(row)]
            else:
                row_dict = {}
                for idx, cell in enumerate(row):
                    header = headers[idx] if idx < len(headers) else f"Column{idx+1}"
                    row_dict[header] = cell
                if any(row_dict.values()):  # Пропускаем пустые строки
                    data.append(row_dict)
        
        return {
            "success": True,
            "rows": len(data),
            "data": data,
            "columns": headers
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при чтении файла: {str(e)}")


@router.post("/import/csv")
async def import_csv(file: UploadFile = File(...)):
    """Импортировать данные из CSV."""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Файл должен быть CSV")
    
    try:
        contents = await file.read()
        # Декодируем с разными кодировками
        try:
            text = contents.decode('utf-8-sig')  # UTF-8 с BOM
        except:
            try:
                text = contents.decode('utf-8')
            except:
                text = contents.decode('cp1251')  # Windows-1251 для русских файлов
        
        csv_reader = csv.DictReader(StringIO(text))
        data = list(csv_reader)
        
        return {
            "success": True,
            "rows": len(data),
            "data": data,
            "columns": list(data[0].keys()) if data else []
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при чтении файла: {str(e)}")


@router.post("/import/json")
async def import_json(file: UploadFile = File(...)):
    """Импортировать данные из JSON."""
    if not file.filename.endswith('.json'):
        raise HTTPException(status_code=400, detail="Файл должен быть JSON")
    
    try:
        contents = await file.read()
        data = json.loads(contents)
        
        return {
            "success": True,
            "data": data
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при чтении файла: {str(e)}")


@router.post("/export/csv")
async def export_csv(data: List[Dict[str, Any]]):
    """Экспортировать данные в CSV."""
    try:
        if not data:
            raise HTTPException(status_code=400, detail="Нет данных для экспорта")
        
        output = StringIO()
        # Получаем заголовки из первого элемента
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
        
        # Конвертируем в bytes с правильной кодировкой
        csv_bytes = output.getvalue().encode('utf-8-sig')
        output_bytes = BytesIO(csv_bytes)
        output_bytes.seek(0)
        
        return StreamingResponse(
            output_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=export.csv"}
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при экспорте: {str(e)}")


@router.post("/export/json")
async def export_json(data: List[Dict[str, Any]]):
    """Экспортировать данные в JSON."""
    try:
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        output = BytesIO(json_str.encode('utf-8'))
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=export.json"}
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при экспорте: {str(e)}")


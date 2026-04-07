"""Сервис для экспорта данных в PDF и Excel."""
from io import BytesIO
from typing import List
from app.models.schemas import DailyPlanItem


class ExportService:
    """Сервис для экспорта отчетов."""
    
    def export_to_excel(
        self, 
        daily_plan: List[DailyPlanItem], 
        total_water: float,
        crop_info: dict = None,
        statistics: dict = None,
        area: float = None,
        region: str = None
    ) -> BytesIO:
        """
        Экспортировать данные в Excel.
        
        Args:
            daily_plan: План по дням
            total_water: Общий объем воды
            
        Returns:
            BytesIO объект с Excel файлом
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        from openpyxl.styles import PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "План водопользования"
        
        # Заголовок отчета
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "ПЛАН ВОДОПОЛЬЗОВАНИЯ ДЛЯ АССОЦИАЦИИ ВОДОПОЛЬЗОВАТЕЛЕЙ"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        ws.row_dimensions[1].height = 30
        
        # Информация о параметрах
        row = 3
        if crop_info:
            ws[f'A{row}'] = "Культура:"
            ws[f'B{row}'] = crop_info.get('name', '')
            row += 1
        if area:
            ws[f'A{row}'] = "Площадь участка:"
            ws[f'B{row}'] = f"{area} га"
            row += 1
        if region:
            ws[f'A{row}'] = "Регион:"
            ws[f'B{row}'] = region
            row += 1
        if statistics:
            ws[f'A{row}'] = "Период планирования:"
            ws[f'B{row}'] = f"{statistics.get('total_days', 0)} дней"
            row += 1
        
        row += 1
        
        # Заголовки таблицы
        headers = ["День", "Объем воды (м³)", "Комментарий", "Дата"]
        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        
        # Стили для заголовков
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Данные
        from datetime import datetime, timedelta
        base_date = datetime.now()
        for item in daily_plan:
            date = base_date + timedelta(days=item.day - 1)
            comment = self._translate_note(item.note)
            ws.append([item.day, item.water, comment, date.strftime("%d.%m.%Y")])
        
        # Итоговая строка
        ws.append([])
        total_row = ws.max_row
        ws.append(["ИТОГО", total_water, "", ""])
        ws[f"B{total_row + 1}"].font = Font(bold=True, size=12)
        ws[f"B{total_row + 1}"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Статистика
        if statistics:
            ws.append([])
            stats_row = ws.max_row + 1
            ws.append(["СТАТИСТИКА", "", "", ""])
            ws[f"A{stats_row}"].font = Font(bold=True)
            ws.append(["Средний объем в день", statistics.get('average_daily', 0), "", ""])
            ws.append(["Максимальный объем", statistics.get('max_daily', 0), "", ""])
            ws.append(["Минимальный объем", statistics.get('min_daily', 0), "", ""])
            ws.append(["Дней с дождем", statistics.get('rain_days', 0), "", ""])
            ws.append(["Дней с высокой температурой", statistics.get('high_temp_days', 0), "", ""])
        
        # Автоширина колонок
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_to_pdf(
        self, 
        daily_plan: List[DailyPlanItem], 
        total_water: float,
        crop_info: dict = None,
        statistics: dict = None,
        area: float = None,
        region: str = None
    ) -> BytesIO:
        """
        Экспортировать данные в PDF.
        
        Args:
            daily_plan: План по дням
            total_water: Общий объем воды
            
        Returns:
            BytesIO объект с PDF файлом
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        
        # Регистрируем шрифт с поддержкой кириллицы
        # Используем стандартный подход reportlab для Unicode
        # В reportlab для кириллицы нужно использовать шрифты с поддержкой Unicode
        # или использовать встроенные шрифты через правильную кодировку
        
        # Пытаемся использовать шрифт, который поддерживает кириллицу
        # В Windows обычно есть Arial, который поддерживает кириллицу
        font_name = 'Helvetica'
        
        # Для правильного отображения кириллицы в reportlab нужно:
        # 1. Использовать шрифт с поддержкой Unicode (через TTF)
        # 2. Или использовать правильную кодировку в Paragraph
        
        # Пробуем зарегистрировать системный шрифт с поддержкой кириллицы
        try:
            # Windows пути к шрифтам
            font_paths = [
                'C:/Windows/Fonts/arial.ttf',
                'C:/Windows/Fonts/arialuni.ttf',  # Arial Unicode MS
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
                '/System/Library/Fonts/Helvetica.ttc',  # macOS
            ]
            
            registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                        font_name = 'CyrillicFont'
                        registered = True
                        break
                    except:
                        continue
            
            if not registered:
                # Если не удалось зарегистрировать, используем стандартный подход
                # В reportlab для кириллицы можно использовать правильную кодировку
                font_name = 'Helvetica'
        except Exception as e:
            # В случае ошибки используем стандартный шрифт
            font_name = 'Helvetica'
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            fontName=font_name,
        )
        
        # Заголовок
        title_text = "ПЛАН ВОДОПОЛЬЗОВАНИЯ ДЛЯ АССОЦИАЦИИ ВОДОПОЛЬЗОВАТЕЛЕЙ"
        # В reportlab Paragraph автоматически обрабатывает Unicode, если используется правильный шрифт
        title = Paragraph(title_text, title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.5*cm))
        
        # Информация о параметрах
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#34495e'),
            fontName=font_name,
        )
        
        if crop_info or area or region:
            info_text = []
            if crop_info:
                info_text.append(f"<b>Культура:</b> {crop_info.get('name', '')}")
            if area:
                info_text.append(f"<b>Площадь участка:</b> {area} га")
            if region:
                info_text.append(f"<b>Регион:</b> {region}")
            if statistics:
                info_text.append(f"<b>Период планирования:</b> {statistics.get('total_days', 0)} дней")
            
            info_para = Paragraph("<br/>".join(info_text), info_style)
            elements.append(info_para)
            elements.append(Spacer(1, 0.3*cm))
        
        # Таблица
        from datetime import datetime, timedelta
        base_date = datetime.now()
        
        # Используем правильную кодировку для всех текстовых данных
        data = [
            ["День", "Объем воды (м³)", "Комментарий", "Дата"]
        ]
        
        for item in daily_plan:
            date = base_date + timedelta(days=item.day - 1)
            comment = self._translate_note(item.note)
            # Убеждаемся, что все строки правильно закодированы
            data.append([
                str(item.day), 
                f"{item.water:.2f}", 
                comment, 
                date.strftime("%d.%m.%Y")
            ])
        
        # Итоговая строка
        data.append(["ИТОГО", f"{total_water:.2f}", "", ""])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F2F2F2')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D0D0D0')),
            ('FONTNAME', (0, -1), (-1, -1), font_name),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
            # Добавляем шрифт для всех ячеек
            ('FONTNAME', (0, 1), (-1, -2), font_name),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Статистика
        if statistics:
            stats_title = Paragraph("Статистика", ParagraphStyle(
                'StatsTitle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#2c3e50'),
                fontName=font_name,
            ))
            elements.append(stats_title)
            elements.append(Spacer(1, 0.2*cm))
            
            stats_data = [
                ["Показатель", "Значение"],
                ["Средний объем в день", f"{statistics.get('average_daily', 0):.2f} м³"],
                ["Максимальный объем", f"{statistics.get('max_daily', 0):.2f} м³"],
                ["Минимальный объем", f"{statistics.get('min_daily', 0):.2f} м³"],
                ["Дней с дождем", str(statistics.get('rain_days', 0))],
                ["Дней с высокой температурой", str(statistics.get('high_temp_days', 0))],
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#95A5A6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECF0F1')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
            ]))
            elements.append(stats_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    def _translate_note(self, note: str) -> str:
        """Перевести код комментария на русский."""
        translations = {
            "normal": "Норма",
            "rain_expected": "Ожидается дождь",
            "high_temperature": "Высокая температура",
        }
        return translations.get(note, note)

